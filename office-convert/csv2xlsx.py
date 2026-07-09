#!/usr/bin/env python3
"""csv2xlsx - Convert CSV file(s) to a styled Excel .xlsx workbook.

Usage:
  csv2xlsx <input.csv> [output.xlsx] [--text] [--no-header]
  csv2xlsx <a.csv> <b.csv> [...] -o <output.xlsx> [--text] [--no-header]

Reads RFC 4180 CSV and writes a workbook with a bold, frozen header row, an
autofilter, and auto-fitted column widths. One CSV makes a single-sheet
workbook; several CSVs make one sheet each, named after the input filename
stem, in the order given. Multiple inputs require -o/--output to name the
workbook.

By default cells are conservatively typed so Excel treats numbers as numbers
and dates as dates -- but only when the conversion is lossless. Values that
Excel famously mangles are kept as text: leading-zero codes (007, zip codes),
phone-style strings (+1 555...), and digit runs longer than 15 (Excel loses
precision past that). Use --text to disable all inference and write every
cell verbatim as text.
"""
import argparse
import csv
import os
import re
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Strings that look numeric but must stay text or Excel corrupts them.
_LEADING_ZERO = re.compile(r"^-?0\d")          # 007, 00123  (but not 0 or 0.5)
_PHONE_ISH = re.compile(r"^\+\d")               # +1 555 1234
_DATE_FORMATS = ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S")

# Characters openpyxl forbids in a sheet title.
_BAD_TITLE = re.compile(r"[\\/*?:\[\]]")


def infer(value):
    """Return a typed value for an xlsx cell, or the original string as text."""
    if value == "":
        return None
    s = value.strip()
    if _LEADING_ZERO.match(s) or _PHONE_ISH.match(s):
        return value
    # Integer, only if it round-trips exactly and won't overflow Excel's
    # 15-significant-digit precision.
    if re.fullmatch(r"-?\d+", s) and len(s.lstrip("-")) <= 15:
        return int(s)
    # Float, only if it parses and is finite.
    try:
        f = float(s)
        if f == f and f not in (float("inf"), float("-inf")):
            return f
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return value


def sheet_title(path, used):
    """A unique, Excel-legal sheet title from an input filename stem."""
    stem = os.path.splitext(os.path.basename(path))[0]
    name = _BAD_TITLE.sub("_", stem)[:31] or "Sheet"
    base, n = name, 2
    while name.lower() in used:
        suffix = f"_{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


def write_sheet(ws, input_path, as_text, no_header):
    """Fill one worksheet from one CSV, with the header/filter/width styling."""
    try:
        f = open(input_path, newline="", encoding="utf-8-sig")
    except OSError as e:
        sys.exit(f"Error: {e}")

    widths = {}
    with f:
        reader = csv.reader(f)
        for r, row in enumerate(reader, start=1):
            for c, raw in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c)
                cell.value = raw if as_text else infer(raw)
                widths[c] = max(widths.get(c, 0), len(raw))
            if r == 1 and not no_header:
                for cell in ws[1]:
                    cell.font = Font(bold=True)

    if ws.max_row == 0:
        sys.exit(f"Error: {input_path} is empty")

    if not no_header:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 8), 60)


def resolve_io(inputs, output, error):
    """Map positionals + -o onto (inputs, output), keeping the legacy
    `csv2xlsx <in> [out]` form working while requiring -o past two files."""
    if output:
        return inputs, output
    if len(inputs) == 1:
        return inputs, re.sub(r"\.[^.]*$", "", inputs[0]) + ".xlsx"
    if len(inputs) == 2:
        return inputs[:1], inputs[1]  # legacy: input, output
    error("with more than two files, use -o/--output to name the workbook")


def main():
    ap = argparse.ArgumentParser(prog="csv2xlsx", description="Convert CSV file(s) to a styled .xlsx workbook.")
    ap.add_argument("inputs", nargs="+", metavar="input", help="input .csv file(s)")
    ap.add_argument("-o", "--output", help="output .xlsx (required for more than two inputs)")
    ap.add_argument("--text", action="store_true", help="write every cell as text; no type inference")
    ap.add_argument("--no-header", action="store_true", help="treat the first row as data, not a header")
    args = ap.parse_args()

    inputs, out = resolve_io(args.inputs, args.output, ap.error)

    wb = Workbook()
    used = set()
    for i, path in enumerate(inputs):
        ws = wb.active if i == 0 else wb.create_sheet()
        if len(inputs) > 1:
            ws.title = sheet_title(path, used)
        write_sheet(ws, path, args.text, args.no_header)

    wb.save(out)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
