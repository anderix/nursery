#!/usr/bin/env python3
"""xlsx2csv - Convert an Excel .xlsx worksheet to RFC 4180 CSV.

Usage: xlsx2csv <input.xlsx> [output.csv] [--sheet NAME] [--list]

The reverse of csv2xlsx. Converts the active sheet by default; use --sheet to
pick another, or --list to print the sheet names and exit. Dates are written
in ISO 8601, and formula cells are written as their last computed value.
"""
import argparse
import csv
import re
import sys
from datetime import date, datetime

from openpyxl import load_workbook


def fmt(value):
    """Render a cell value as a CSV string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Drop a midnight time component so date-only cells stay clean.
        if value.hour == value.minute == value.second == 0:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def main():
    ap = argparse.ArgumentParser(prog="xlsx2csv", description="Convert an .xlsx worksheet to RFC 4180 CSV.")
    ap.add_argument("input", help="input .xlsx file")
    ap.add_argument("output", nargs="?", help="output .csv (default: input name with .csv)")
    ap.add_argument("--sheet", help="worksheet name (default: the active sheet)")
    ap.add_argument("--list", action="store_true", help="list worksheet names and exit")
    args = ap.parse_args()

    try:
        wb = load_workbook(args.input, read_only=True, data_only=True)
    except OSError as e:
        sys.exit(f"Error: {e}")

    if args.list:
        print("\n".join(wb.sheetnames))
        return

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            sys.exit(f"Error: no sheet named {args.sheet!r}. Available: {', '.join(wb.sheetnames)}")
        ws = wb[args.sheet]
    else:
        ws = wb.active

    out = args.output or re.sub(r"\.[^.]*$", "", args.input) + ".csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, lineterminator="\n")
        for row in ws.iter_rows(values_only=True):
            writer.writerow([fmt(v) for v in row])

    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
